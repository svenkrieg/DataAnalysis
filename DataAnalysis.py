import re
import zipfile
import statistics
from io import TextIOWrapper
from datetime import datetime
from collections import defaultdict
import tkinter as tk
from tkinter import filedialog, messagebox

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def count_words(text):
    if not text:
        return 0
    return len(re.findall(r"\b\w+\b", text, flags=re.UNICODE))


def sanitize_sheet_name(name):
    if not name:
        name = "Unbekannt"
    name = re.sub(r'[\\/*?:\[\]]', "_", str(name))
    return name[:31]


def unique_sheet_name(wb, base_name):
    existing = set(ws.title for ws in wb.worksheets)
    if base_name not in existing:
        return base_name

    counter = 1
    while True:
        suffix = f"_{counter}"
        shortened = base_name[:31 - len(suffix)] + suffix
        if shortened not in existing:
            return shortened
        counter += 1


def parse_log_stream(stream):
    entries = []
    current = {}
    content_lines = []
    collecting_content = False

    for raw_line in stream:
        line = raw_line.rstrip("\n")

        if line.startswith("TYPE:"):
            if current:
                if content_lines:
                    current["CONTENT"] = "\n".join(content_lines).strip()
                else:
                    current["CONTENT"] = current.get("CONTENT", "")
                entries.append(current)

            current = {"TYPE": line.split(": ", 1)[1].strip()}
            content_lines = []
            collecting_content = False
            continue

        if collecting_content and not re.match(r"^[A-Z_]+:\s", line):
            content_lines.append(line)
            continue

        if line.startswith("TIME:"):
            time_str = line.split(": ", 1)[1].strip()
            current["TIME"] = datetime.fromisoformat(time_str.replace("Z", "+00:00"))
            collecting_content = False

        elif line.startswith("URL:"):
            current["URL"] = line.split(": ", 1)[1].strip()
            collecting_content = False

        elif line.startswith("USER:"):
            current["USER"] = line.split(": ", 1)[1].strip()
            collecting_content = False

        elif line.startswith("THREAD_ID:"):
            current["THREAD_ID"] = line.split(": ", 1)[1].strip()
            collecting_content = False

        elif line.startswith("CONTENT:"):
            first_content = line.split(": ", 1)[1]
            content_lines = [first_content] if first_content else []
            collecting_content = True

    if current:
        if content_lines:
            current["CONTENT"] = "\n".join(content_lines).strip()
        else:
            current["CONTENT"] = current.get("CONTENT", "")
        entries.append(current)

    return entries


def analyze_entries(entries):
    stats = defaultdict(lambda: {
        "user_messages": 0,
        "user_words": 0,
        "assistant_messages": 0,
        "assistant_words": 0,
        "times": []
    })

    for e in entries:
        url = e.get("URL", "UNKNOWN")
        event_type = e.get("TYPE", "")
        content = e.get("CONTENT", "")
        time = e.get("TIME")

        if time is not None:
            stats[url]["times"].append(time)

        words = count_words(content)

        if event_type == "USER_MESSAGE":
            stats[url]["user_messages"] += 1
            stats[url]["user_words"] += words

        elif event_type == "ASSISTANT_MESSAGE":
            stats[url]["assistant_messages"] += 1
            stats[url]["assistant_words"] += words

    rows = []

    total_user_messages = 0
    total_user_words = 0
    total_assistant_messages = 0
    total_assistant_words = 0
    total_turns = 0
    total_duration_seconds = 0.0

    for url in sorted(stats.keys()):
        data = stats[url]

        if data["times"]:
            duration_seconds = (max(data["times"]) - min(data["times"])).total_seconds()
        else:
            duration_seconds = 0.0

        turns = min(data["user_messages"], data["assistant_messages"])

        row = {
            "URL": url,
            "User_Messages": data["user_messages"],
            "User_Words": data["user_words"],
            "Assistant_Messages": data["assistant_messages"],
            "Assistant_Words": data["assistant_words"],
            "Turns": turns,
            "Duration_Seconds": round(duration_seconds, 2)
        }
        rows.append(row)

        total_user_messages += data["user_messages"]
        total_user_words += data["user_words"]
        total_assistant_messages += data["assistant_messages"]
        total_assistant_words += data["assistant_words"]
        total_turns += turns
        total_duration_seconds += duration_seconds

    total_row = {
        "URL": "TOTAL",
        "User_Messages": total_user_messages,
        "User_Words": total_user_words,
        "Assistant_Messages": total_assistant_messages,
        "Assistant_Words": total_assistant_words,
        "Turns": total_turns,
        "Duration_Seconds": round(total_duration_seconds, 2)
    }

    rows.append(total_row)
    return rows, total_row


def autosize_columns(ws):
    for column_cells in ws.columns:
        max_length = 0
        col_idx = column_cells[0].column
        col_letter = get_column_letter(col_idx)

        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))

        ws.column_dimensions[col_letter].width = min(max_length + 2, 80)


def style_header(row):
    fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)

    for cell in row:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def style_total_row(row):
    fill = PatternFill(fill_type="solid", fgColor="D9EAD3")
    font = Font(bold=True)

    for cell in row:
        cell.fill = fill
        cell.font = font


def style_section_header(row):
    fill = PatternFill(fill_type="solid", fgColor="808080")
    font = Font(color="FFFFFF", bold=True)

    for cell in row:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def extract_username(entries, fallback_filename):
    for entry in entries:
        user = entry.get("USER", "").strip()
        if user:
            return user

    fallback = fallback_filename.rsplit("/", 1)[-1]
    fallback = fallback.rsplit("\\", 1)[-1]
    fallback = re.sub(r"\.txt$", "", fallback, flags=re.IGNORECASE)
    return fallback or "Unbekannt"


def add_summary_sheet(wb, username, summary_rows):
    headers = [
        "URL",
        "User_Messages",
        "User_Words",
        "Assistant_Messages",
        "Assistant_Words",
        "Turns",
        "Duration_Seconds"
    ]

    base_name = sanitize_sheet_name(f"{username}_Auswertung")
    sheet_name = unique_sheet_name(wb, base_name)
    ws = wb.create_sheet(title=sheet_name)

    ws.append(headers)

    for row in summary_rows:
        ws.append([row[h] for h in headers])

    style_header(ws[1])
    style_total_row(ws[ws.max_row])

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top")

    autosize_columns(ws)
    ws.freeze_panes = "A2"


def add_rawdata_sheet(wb, username, entries):
    headers = ["USER", "TYPE", "TIME", "URL", "THREAD_ID", "CONTENT"]

    base_name = sanitize_sheet_name(f"{username}_Rohdaten")
    sheet_name = unique_sheet_name(wb, base_name)
    ws = wb.create_sheet(title=sheet_name)

    ws.append(headers)

    for entry in entries:
        ws.append([
            entry.get("USER", ""),
            entry.get("TYPE", ""),
            entry.get("TIME", "").isoformat() if entry.get("TIME") else "",
            entry.get("URL", ""),
            entry.get("THREAD_ID", ""),
            entry.get("CONTENT", "")
        ])

    style_header(ws[1])

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    autosize_columns(ws)
    ws.column_dimensions["F"].width = min(max(ws.column_dimensions["F"].width, 40), 100)
    ws.freeze_panes = "A2"


def safe_mean(values):
    return round(statistics.mean(values), 2) if values else ""


def safe_median(values):
    return round(statistics.median(values), 2) if values else ""


def safe_min(values):
    return round(min(values), 2) if values else ""


def safe_max(values):
    return round(max(values), 2) if values else ""


def safe_stdev(values):
    return round(statistics.stdev(values), 2) if len(values) > 1 else 0


def add_overall_sheet(wb, participant_totals):
    ws = wb.active
    ws.title = "Gesamtauswertung"

    participant_headers = [
        "USER",
        "User_Messages",
        "User_Words",
        "Assistant_Messages",
        "Assistant_Words",
        "Turns",
        "Duration_Seconds"
    ]

    ws.append(["Gesamtauswertung über alle VP"])
    ws.append([])
    ws.append(participant_headers)

    for row in participant_totals:
        ws.append([
            row["USER"],
            row["User_Messages"],
            row["User_Words"],
            row["Assistant_Messages"],
            row["Assistant_Words"],
            row["Turns"],
            row["Duration_Seconds"]
        ])

    style_section_header(ws[1])
    style_header(ws[3])

    start_stats_row = ws.max_row + 2
    ws.append(["Statistische Kennwerte über alle VP"])
    style_section_header(ws[ws.max_row])

    stats_headers = ["Kennwert"] + participant_headers[1:]
    ws.append(stats_headers)
    style_header(ws[ws.max_row])

    user_messages_vals = [r["User_Messages"] for r in participant_totals]
    user_words_vals = [r["User_Words"] for r in participant_totals]
    assistant_messages_vals = [r["Assistant_Messages"] for r in participant_totals]
    assistant_words_vals = [r["Assistant_Words"] for r in participant_totals]
    turns_vals = [r["Turns"] for r in participant_totals]
    duration_vals = [r["Duration_Seconds"] for r in participant_totals]

    stat_rows = [
        ["N",
         len(participant_totals),
         len(participant_totals),
         len(participant_totals),
         len(participant_totals),
         len(participant_totals),
         len(participant_totals)],

        ["Mittelwert",
         safe_mean(user_messages_vals),
         safe_mean(user_words_vals),
         safe_mean(assistant_messages_vals),
         safe_mean(assistant_words_vals),
         safe_mean(turns_vals),
         safe_mean(duration_vals)],

        ["Median",
         safe_median(user_messages_vals),
         safe_median(user_words_vals),
         safe_median(assistant_messages_vals),
         safe_median(assistant_words_vals),
         safe_median(turns_vals),
         safe_median(duration_vals)],

        ["Minimum",
         safe_min(user_messages_vals),
         safe_min(user_words_vals),
         safe_min(assistant_messages_vals),
         safe_min(assistant_words_vals),
         safe_min(turns_vals),
         safe_min(duration_vals)],

        ["Maximum",
         safe_max(user_messages_vals),
         safe_max(user_words_vals),
         safe_max(assistant_messages_vals),
         safe_max(assistant_words_vals),
         safe_max(turns_vals),
         safe_max(duration_vals)],

        ["Standardabweichung",
         safe_stdev(user_messages_vals),
         safe_stdev(user_words_vals),
         safe_stdev(assistant_messages_vals),
         safe_stdev(assistant_words_vals),
         safe_stdev(turns_vals),
         safe_stdev(duration_vals)]
    ]

    for row in stat_rows:
        ws.append(row)

    for row in ws.iter_rows(min_row=1):
        for cell in row:
            cell.alignment = Alignment(vertical="top")

    autosize_columns(ws)
    ws.freeze_panes = "A3"


def choose_zip_file():
    root = tk.Tk()
    root.withdraw()
    root.update()

    zip_path = filedialog.askopenfilename(
        title="ZIP-Datei mit Logdateien auswählen",
        filetypes=[("ZIP-Dateien", "*.zip")]
    )

    root.destroy()
    return zip_path


def choose_output_file():
    root = tk.Tk()
    root.withdraw()
    root.update()

    output_path = filedialog.asksaveasfilename(
        title="Excel-Datei speichern unter",
        defaultextension=".xlsx",
        filetypes=[("Excel-Dateien", "*.xlsx")],
        initialfile="chat_analysis.xlsx"
    )

    root.destroy()
    return output_path


def process_zip_to_excel(zip_path, output_file):
    wb = Workbook()

    with zipfile.ZipFile(zip_path, "r") as zf:
        txt_files = sorted([name for name in zf.namelist() if name.lower().endswith(".txt")])

        if not txt_files:
            raise ValueError("In der ZIP-Datei wurden keine .txt-Dateien gefunden.")

        processed = []
        participant_totals = []

        for txt_name in txt_files:
            with zf.open(txt_name, "r") as raw_file:
                stream = TextIOWrapper(raw_file, encoding="utf-8")
                entries = parse_log_stream(stream)

            if not entries:
                continue

            username = extract_username(entries, txt_name)
            summary_rows, total_row = analyze_entries(entries)

            participant_totals.append({
                "USER": username,
                "User_Messages": total_row["User_Messages"],
                "User_Words": total_row["User_Words"],
                "Assistant_Messages": total_row["Assistant_Messages"],
                "Assistant_Words": total_row["Assistant_Words"],
                "Turns": total_row["Turns"],
                "Duration_Seconds": total_row["Duration_Seconds"]
            })

            processed.append({
                "username": username,
                "entries": entries,
                "summary_rows": summary_rows
            })

        if not processed:
            raise ValueError("Es konnten keine verwertbaren Logdateien verarbeitet werden.")

        add_overall_sheet(wb, participant_totals)

        for item in processed:
            add_summary_sheet(wb, item["username"], item["summary_rows"])

        for item in processed:
            add_rawdata_sheet(wb, item["username"], item["entries"])

    wb.save(output_file)
    return len(processed)


def main():
    zip_path = choose_zip_file()
    if not zip_path:
        print("Keine ZIP-Datei ausgewählt.")
        return

    output_path = choose_output_file()
    if not output_path:
        print("Kein Speicherort ausgewählt.")
        return

    try:
        count = process_zip_to_excel(zip_path, output_path)
        print(f"Fertig. {count} TXT-Datei(en) verarbeitet.")
        print(f"Excel-Datei erstellt: {output_path}")

        root = tk.Tk()
        root.withdraw()
        messagebox.showinfo(
            "Erfolg",
            f"{count} TXT-Datei(en) verarbeitet.\n\nExcel-Datei erstellt:\n{output_path}"
        )
        root.destroy()

    except Exception as e:
        print(f"Fehler: {e}")

        root = tk.Tk()
        root.withdraw()
        messagebox.showerror("Fehler", str(e))
        root.destroy()


if __name__ == "__main__":
    main()